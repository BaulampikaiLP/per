from openpyxl import load_workbook


def fetch_login_data(file_path):
    workbook = load_workbook(file_path)
    sheet = workbook.active
    data = []

    for row in sheet.iter_rows(min_row=5, values_only=True):
        data.append(row)

    return data
