import openpyxl
from openpyxl.reader.excel import load_workbook

try:
    wb = openpyxl.load_workbook("D:\\demo\\read.xlsx")
    sh1 = wb['Sheet1']
    sh1['A1'] = 'hello'
    sh1['B2'] = 'xyz'
    wb.save("D:\\demo\\read.xlsx")
    for row in sh1.iter_rows(values_only=True):
        print(row)

except FileNotFoundError:
    print("file not found")
