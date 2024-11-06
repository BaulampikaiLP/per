from openpyxl import Workbook


# Class One
class One:
    def __init__(self):
        # Create an Excel file with 5 entries of names and ages
        self.data = [
            {"name": "John", "age": 25},
            {"name": "Alice", "age": 30},
            {"name": "Bob", "age": 22},
            {"name": "Emma", "age": 28},
            {"name": "Charlie", "age": 26}
        ]

    def create_excel(self):
        try:
            # Create a workbook and add data
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Names and Ages"
            sheet.append(["Name", "Age"])  # Header row

            for entry in self.data:
                sheet.append([entry["name"] + "_A", entry["age"]])

            # Save the workbook
            workbook.save('D:\\Names and Ages.xlsx')
            print("Excel file created with names and ages.")
        except Exception as e:
            print(f"Error while creating Excel file: {e}")

    def separate_data(self):
        try:
            # Separate names and ages into different Excel files
            name_workbook = Workbook()
            age_workbook = Workbook()

            name_sheet = name_workbook.active
            age_sheet = age_workbook.active

            name_sheet.title = "Name-n"
            age_sheet.title = "Ages"

            name_sheet.append(["Name"])  # Header row
            age_sheet.append(["Age"])  # Header row

            # Mimic do-while using a while loop
            i = 0
            while True:
                # Get the name and age
                name = self.data[i]["name"]
                age = self.data[i]["age"]

                # Store name and age in different Excel files
                if isinstance(name, str):
                    name_sheet.append([name])
                if isinstance(age, int):
                    age_sheet.append([age])

                # Increment and check condition to break (this simulates the 'do-while' loop)
                i += 1
                if i >= len(self.data):
                    break

            # Save the workbooks
            name_workbook.save("names.xlsx")
            age_workbook.save("ages.xlsx")
            print("Names and ages separated into different Excel files.")
        except Exception as e:
            print(f"Error while separating data: {e}")


# Class Two
class Two:
    def __init__(self):
        pass

    def process(self):
        # Create an object of class One and call its methods
        one_obj = One()
        one_obj.create_excel()
        one_obj.separate_data()


# Driver code
if __name__ == "__main__":
    two_obj = Two()
    two_obj.process()
